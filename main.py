# Import dependencies
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import yfinance as yf
import datetime as dt
import mplfinance as mpf
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import LineChart, Reference
from scipy import stats
from matplotlib.widgets import RangeSlider
import matplotlib.dates as mdates


############################  PERIOD  ########################################




# input the symbol for data retrieval
symbol = input("Enter stock symbol (e.g., RELIANCE.NS): ")
start = dt.date.today() - dt.timedelta(days=365 * 2)
end = dt.date.today()

# Read data
df = yf.download(symbol, period="2y")

if df.empty:
    print(f"Invalid symbol or no data found for: {symbol}")
    exit()

df.columns = df.columns.get_level_values(0)                     #tO GET PROPER HEADERS FOR OUR DATAFRAME
df = df.reset_index()                                           #RESET INDEX TO MAKE DATE A COLUMN
df.rename(columns={df.columns[0]: "Date"}, inplace=True)        #rENAME DATA COLUMN
df["Date"] = pd.to_datetime(df["Date"]).dt.date                 #rEMOVE TIME FROM DATE COLUMN


high = df["High"].squeeze()
low = df["Low"].squeeze()
close = df["Close"].squeeze()
volume = df["Volume"].squeeze()
open = df["Open"].squeeze()


# Save data into a xlsx file new sheet
excel_path = "stocks.xlsx"
with pd.ExcelWriter(
    excel_path, 
    engine="openpyxl",
    mode="a", 
    if_sheet_exists="replace"
) as writer:
    df.to_excel(writer, sheet_name=symbol, index=False)




###############################  Small tweak for excel  #####################################


#Sheet names update for the dashboard
excel_path = "stocks.xlsx"

wb = load_workbook(excel_path)

# Get stock sheets only (exclude Dashboard & helper)
stock_sheets = [
    s for s in wb.sheetnames
    if not s.startswith("__") and s != "Dashboard" and s != "__SHEETS__"      #Avoid junk sheets  Dashboard and helper
]


# Write sheet names to helper sheet
with pd.ExcelWriter(
    excel_path,
    engine="openpyxl",
    mode="a",
    if_sheet_exists="replace"
) as writer:
    pd.DataFrame(stock_sheets, columns=["Symbol"]) \
      .to_excel(writer, sheet_name="__SHEETS__", index=False)
    
stock_sheets.sort()                                     #Sort the sheet names


## Save workbook

wb.save(excel_path)



######################################  PERIOD  ########################################


########## Statistical Risk & Return Analysis Using Probability Distributions ##########


######################################  PERIOD  ########################################



excel_path = "stocks.xlsx"

df = pd.read_excel(excel_path, sheet_name=symbol)   ## Load data for selected stock symbol          
df["Date"] = pd.to_datetime(df["Date"])             ## Ensure Date column is datetime
df.sort_values("Date", inplace=True)                ## Ensure data is sorted by date

df["date_num"] = pd.to_datetime(df["Date"]).map(pd.Timestamp.toordinal)  ## Convert dates to numerical format for analysis

date_min = df["date_num"].min()             ## Get min and max date numbers for reference   
date_max = df["date_num"].max()                         



## Log Returns Calculation

df["log_return"] = np.log(df["Close"] / df["Close"].shift(1))   #Log return calculation using numpy log function    
returns = df["log_return"].dropna()                                             #Drop NA values from log returns

stats_summary = {
    "Mean Return": returns.mean(),              #The average daily return
    "Volatility (Std Dev)": returns.std(),      #The standard deviation of returns
    "Skewness": returns.skew(),          #Measure of asymmetry of the return distribution
    "Kurtosis": returns.kurtosis()         #Measure of the "tailedness" of the return distribution
}

for k, v in stats_summary.items():           #Print the statistical summary 
    print(f"{k}: {v:.6f}")                     #Formatted to 6 decimal places     


############################  PERIOD  ######################################## 

## wE ARE USING RETURNS VARIABLE AS OUR RANDOM VARIABLE FOR FITTING DISTRIBUTIONS AND RISK METRICS CALCULATION. ##
## THIS VARIABLE CONTAINS THE LOG RETURNS OF THE STOCK PRICES. ALL SUBSEQUENT ANALYSES WILL BE BASED ON THIS RETURNS DATA.##

############################  PERIOD  ########################################


# normal distribution

mu, sigma = stats.norm.fit(returns)                                         #Fit a normal distribution to the returns data

x = np.linspace(returns.min(), returns.max(), 500)                          #Generate x values for plotting the fitted distribution
pdf_normal = stats.norm.pdf(x, mu, sigma)                      #Calculate the PDF of the fitted normal distribution                 




## sTUDENT'S t DISTRIBUTION 

params = stats.t.fit(returns)                       #Fit a Student's t-distribution to the returns data         
pdf_t = stats.t.pdf(x, *params)                      #Calculate the PDF of the fitted Student's t-distribution                      




################# Monte Carlo simulation ################

sim_days = 252                                                          # Number of trading days in a year
simulations = 10000                                                     # Number of simulation paths

sim_returns = np.random.normal(mu, sigma, (simulations, sim_days))      # Generate random returns based on fitted normal distribution
sim_price_paths = np.exp(sim_returns.cumsum(axis=1))                    # Simulate price paths


################  Risk Metrics Calculation  ##############


# Value at Risk (VaR) Calculation

confidence = 0.95                               # Confidence level for VaR
VaR = np.percentile(returns, (1 - confidence) * 100)    # Calculate VaR at the specified confidence level
print(f"95% Value at Risk: {VaR:.4%}")          


# Conditional Value at Risk (CVaR) Calculation

CVaR = returns[returns <= VaR].mean()           # Calculate CVaR as the average loss beyond the VaR threshold
print(f"95% CVaR: {CVaR:.4%}")                  


# Probability of daily loss greater than 2%

prob_loss = np.mean(returns < -0.02)                        # Calculate the probability of daily loss exceeding 2%
print(f"Probability of >2% daily loss: {prob_loss:.2%}")


############################  PERIOD  ########################################




########################## PLOTTING PART USING MATPLOT #############################

# ===================== DASHBOARD PLOT =====================
fig, axes = plt.subplots(3, 2, figsize=(16, 15))                                                    # Create a 3x2 grid of subplots
plt.subplots_adjust(bottom=0.15)                                                            # Adjust bottom to make space for slider

fig.suptitle(f"Statistical Risk & Return Analysis — {symbol}", fontsize=18)                         # Overall title for the dashboard

slider_ax = plt.axes([0.15, 0.05, 0.7, 0.03])                                   # Axes for the date range slider



############## Adding a date range slider #######################

date_slider = RangeSlider(                                                          
    slider_ax,
    "Date Range",
    date_min,
    date_max,
    valinit=(date_min, date_max),
)


############ Helper function to convert ordinal back to date ###############

def ordinal_to_date(x):                                                                             
    if isinstance(x, (int, float)):                                                     # Check if x is ordinal number
        return pd.Timestamp.fromordinal(int(x)).strftime("%Y-%m-%d")                    # Convert to date string
    elif isinstance(x, (pd.Timestamp, dt.date)):                                        # Check if x is already a date
        return pd.to_datetime(x).strftime("%Y-%m-%d")                               # Format date as string 
    else:                                                               # Fallback for unexpected types             
        return str(x)                                                           



########### Initial data setup for plotting ##################


fig.suptitle(                                                                               # Overall title for the dashboard
    f"Statistical Risk & Return Analysis — {symbol}\n"
    f"Period: {ordinal_to_date(start)} → {ordinal_to_date(end)}",
    fontsize=18
)


############# Update function for slider interaction #######################


def update(val):                                                                # Update plots based on selected date range
    start, end = date_slider.val
    start, end = int(start), int(end)

    # Update slider label with readable dates
    date_slider.valtext.set_text(
    f"{ordinal_to_date(start)}  →  {ordinal_to_date(end)}"
    )
    date_slider.label.set_text("Analysis Date Range")



    dff = df[(df["date_num"] >= start) & (df["date_num"] <= end)].copy()                #Filter data based on selected date range

    if len(dff) < 30:
        return


    # Recompute returns

    dff["log_return"] = np.log(dff["Close"] / dff["Close"].shift(1))
    returns = dff["log_return"].dropna()


    # Risk metrics

    mu, sigma = stats.norm.fit(returns)
    params = stats.t.fit(returns)

    VaR = np.percentile(returns, 5)
    CVaR = returns[returns <= VaR].mean()
    prob_loss = np.mean(returns < -0.02)


    #################### distribution stats  ############################


    skew = returns.skew()                                                       #Skewness of returns
    kurt = returns.kurtosis()                                               #Kurtosis of returns                            


    ########## x values for plotting ##############


    VaR = np.percentile(returns, 5)                                                 #Calculate VaR at the specified confidence level
    CVaR = returns[returns <= VaR].mean()                                           #Calculate CVaR as the average loss beyond the VaR threshold
    prob_loss = np.mean(returns < -0.02)                                        #Calculate the probability of daily loss exceeding 2%

    x = np.linspace(returns.min(), returns.max(), 500)                                  #Generate x values for plotting the fitted distribution








########################################  Risk Interpretation Text  ###########################################################



    if VaR > -0.01:                                         
        var_text = "Very low short-term downside risk"
    elif -0.03 < VaR <= -0.01:
        var_text = "Moderate downside risk"
    else:
        var_text = "High downside risk on adverse days"

    if CVaR > -0.02:
        cvar_text = "Losses remain contained even in worst cases"                                                   
    elif -0.05 < CVaR <= -0.02:
        cvar_text = "Significant tail losses during stress periods"
    else:
        cvar_text = "Severe crash risk in extreme market conditions"


    if prob_loss < 0.03:
        prob_text = "Large losses are infrequent"
    elif 0.03 <= prob_loss < 0.07:
        prob_text = "Occasional large losses observed"
    else:
        prob_text = "Frequent large downside moves"

    if skew < -0.5:
        skew_text = "Strong downside asymmetry"
    elif -0.5 <= skew < -0.1:
        skew_text = "Moderate downside bias"
    elif -0.1 <= skew <= 0.1:
        skew_text = "Approximately symmetric returns"
    elif 0.1 < skew <= 0.5:
        skew_text = "Moderate upside bias"
    else:
        skew_text = "Strong upside asymmetry"

    if kurt > 3:
        kurt_text = "Extreme fat tails with high crash probability"
    elif 1 < kurt <= 3:
        kurt_text = "Fat tails present (non-normal extremes)"
    elif -1 <= kurt <= 1:
        kurt_text = "Near-normal tail behavior"
    else:
        kurt_text = "Thin tails with limited extreme risk"




    if kurt > 1 or abs(skew) > 0.5:
        model_text = "Student-t distribution provides a superior fit for tail risk"
    else:
        model_text = "Normal distribution provides an adequate approximation"



    if CVaR < -0.05 or prob_loss > 0.07:
        regime_text = "High-risk regime with elevated tail risk"
    elif CVaR < -0.02 or prob_loss > 0.03:
        regime_text = "Moderate-risk regime with episodic stress"
    else:
        regime_text = "Low-risk regime with stable return behavior"






########################################################### Explanation Text ###########################################################

    
    explanation_text = (                                                                        # Explanation text summarizing risk insights 
        "RISK INTERPRETATION\n"
        "--------------------\n"
        f"• VaR (95%) = {VaR:.2%} → {var_text}.\n"
        f"• CVaR (95%) = {CVaR:.2%} → {cvar_text}.\n"
        f"• P(Loss > 2%) = {prob_loss:.2%} → {prob_text}.\n\n"

        "DISTRIBUTION INSIGHTS\n"
        "--------------------\n"
        f"• Skewness = {skew:.2f} → {skew_text}.\n"
        f"• Kurtosis = {kurt:.2f} → {kurt_text}.\n\n"

        "MODEL COMPARISON\n"
        "--------------------\n"
        f"• {model_text}.\n\n"

        "OVERALL CONCLUSION\n"
        "--------------------\n"
        f"• Current market state: {regime_text}.\n"
        "• Risk characteristics change meaningfully across\n"
        "  different time windows and market conditions."
    )







    ##  Clear axes


    for ax_row in axes:
        for ax in ax_row:
            ax.clear()

    ###################### Plots ######################


    #### Return distribution with normal fit and VaR/CVaR lines #######


    axes[0, 0].hist(
    returns,
    bins=50,
    density=True,                                           # Normalize histogram
    alpha=0.5,
    label="Empirical Returns"
    )   

    axes[0, 0].plot(
    x,
    stats.norm.pdf(x, mu, sigma),
    linewidth=2,                                            # Plot normal PDF
    label="Normal PDF"
    )

    axes[0, 0].axvline(
    VaR,
    linestyle="--",                                         # Plot VaR line
    linewidth=2,
    label="VaR (95%)"
    )

    axes[0, 0].axvline(
        CVaR,
        linestyle=":",                                      # Plot CVaR line
        linewidth=2,
        label="CVaR (95%)"
    )

    ## Set titles and labels

    axes[0, 0].set_title("Return Distribution vs Normal Model")
    axes[0, 0].set_xlabel("Daily Log Returns")
    axes[0, 0].set_ylabel("Probability Density")
    axes[0, 0].legend(fontsize=9)


    ####### Return distribution with Student-t fit and VaR/CVaR lines ####### 

    axes[0, 1].hist(
    returns,
    bins=50,                                                                        # Histogram of returns
    density=True,
    alpha=0.5,
    label="Empirical Returns"
    )

    axes[0, 1].plot(
    x,
    stats.t.pdf(x, *params),                                                                # Plot Student-t PDF
    linewidth=2,
    label="Student-t PDF"
    )

    ## Set titles and labels ##

    axes[0, 1].set_title("Return Distribution vs Student-t Model")
    axes[0, 1].set_xlabel("Daily Log Returns")
    axes[0, 1].set_ylabel("Probability Density")
    axes[0, 1].legend(fontsize=9)

    ######## Q-Q Plots for Model Diagnostics #######

    stats.probplot(returns, dist="norm", plot=axes[1, 0])
    axes[1, 0].set_title("Normal Q-Q Plot (Model Diagnostic)")


    ######## Plot for Student-t Q-Q Plot #######


    stats.probplot(returns, dist=stats.t, sparams=(params[0],), plot=axes[1, 1])
    axes[1, 1].set_title("Student-t Q–Q Plot (Model Diagnostic)")
    axes[2, 0].set_xlim(0, sim_days)

    axes[2, 0].plot(sim_price_paths[:100].T, alpha=0.2)                                     # Plot first 100 simulated price paths

    axes[2, 0].set_title("Monte Carlo Simulated Price Paths")
    axes[2, 0].set_xlabel("Trading Days")                                                   # Monte Carlo simulation plot labels
    axes[2, 0].set_ylabel("Normalized Price Level")


    ########## Interpretation Panel ##########

    axes[2, 1].set_title("Risk Interpretation Summary", loc="left")

    axes[2, 1].axis("off")

    axes[2, 1].text(
        0.02, 0.98,
        explanation_text,
        fontsize=11,
        verticalalignment="top",
        horizontalalignment="left",
        bbox=dict(boxstyle="round", alpha=0.15)
    )


    fig.canvas.draw_idle()


############################  PERIOD  ########################################



date_slider.on_changed(update)                                                  # Call update function on slider change

# Initial draw
update(None)                                                                    # Initial plot rendering

plt.show()                                                                  # Show the dashboard plot

                                                                       





############################  PERIOD  ########################################



#Final messages

print(f"Data for {symbol} saved to {excel_path} successfully.")
print("Available stock sheets updated in __SHEETS__ sheet.")

