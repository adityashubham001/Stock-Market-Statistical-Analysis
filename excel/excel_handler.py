from openpyxl import load_workbook
import pandas as pd

EXCEL_PATH = "stocks.xlsx"


def save_stock(df: pd.DataFrame, symbol: str):                                      # Function to save data in excel
    df = df.copy()                                                                  # Make a copy            

    # Ensure Date is date-only for Excel
    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"]).dt.date

    # Add returns if not present
    if "Returns" not in df.columns:
        df["Returns"] = df["Close"].pct_change()

    with pd.ExcelWriter(                                                            # Write to Excel
        EXCEL_PATH,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace"
    ) as writer:
        df.to_excel(writer, sheet_name=symbol, index=False)                         



def update_sheet_list():                                                            # Function to update sheet list
    wb = load_workbook(EXCEL_PATH)                                                  # Load Excel file    

    stock_sheets = [                                                                 # Get sheet names       
        s for s in wb.sheetnames                                                    # Skip sheets that start with "__"
        if not s.startswith("__") and s != "Dashboard"                              # Skip "Dashboard"
    ]                                                                               # Skip "__SHEETS__"

    stock_sheets.sort()                                                             # Sort    

    with pd.ExcelWriter(                                                            # Write to Excel
        EXCEL_PATH,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace"
    ) as writer:
        pd.DataFrame(stock_sheets, columns=["Symbol"]) \
            .to_excel(writer, sheet_name="__SHEETS__", index=False)

    try:                                                                            # Try to save
        wb.save(EXCEL_PATH)                                                         # Save        
    except PermissionError:                                                         # Exception handling    
        print("⚠ Close the Excel file before running the script.")                     

